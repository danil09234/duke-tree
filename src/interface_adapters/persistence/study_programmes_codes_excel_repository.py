from pathlib import Path
from typing import Any, Generator

import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from src.application.interfaces import Fetchable
from src.interface_adapters.exceptions import InvalidExcelFileStructure


class StudyProgrammesCodesExcelRepository(Fetchable[str]):
    def __init__(self, file_path: Path):
        self._file_path = file_path

    async def fetch_all(self) -> list[str]:
        """
        Fetches study programmes codes from an Excel file.

        :return: List of study programmes codes.
        """
        worksheet = self._get_worksheet()
        programme_codes = self._get_list_of_codes(worksheet)
        return programme_codes

    def _get_worksheet(self) -> Worksheet:
        workbook = openpyxl.load_workbook(self._file_path)
        worksheet: Worksheet = workbook.active
        return worksheet

    @classmethod
    def _get_list_of_codes(cls, worksheet: Worksheet) -> list[str]:
        codes_cells_generator = cls._get_codes_cells_generator(worksheet)
        return [str(cell.value) for cell in codes_cells_generator]

    @classmethod
    def _get_codes_cells_generator(cls, worksheet: Worksheet) -> Generator[Any, None, None]:
        for current_row in cls._get_rows_range(worksheet):
            current_cell = cls._get_code_cell(worksheet, current_row)
            cls._assert_is_valid_code_cell(current_cell)
            yield current_cell

    @classmethod
    def _get_rows_range(cls, worksheet: Worksheet) -> range:
        has_header = cls._sheet_has_header(worksheet)
        first_row_number = 2 if has_header else 1
        last_row_number = worksheet.max_row + 1
        return range(first_row_number, last_row_number)

    @staticmethod
    def _sheet_has_header(worksheet: Worksheet) -> bool:
        header_row = worksheet[1]
        header_cells_values = [cell.value for cell in header_row]
        has_header = any(isinstance(cell, str) for cell in header_cells_values)
        return has_header

    @staticmethod
    def _get_code_cell(worksheet: Worksheet, row: int) -> Cell:
        cell: Cell = worksheet.cell(row=row, column=1)
        return cell

    @staticmethod
    def _assert_is_valid_code_cell(code_cell: Cell) -> None:
        last_row = code_cell.parent.max_row
        if not code_cell.value and code_cell.row != last_row:
            raise InvalidExcelFileStructure(f"Cell A{code_cell} is empty")
